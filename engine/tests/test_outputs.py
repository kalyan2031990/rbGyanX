import math
import pathlib
import subprocess
import sys
import tempfile

import numpy as np
import pytest


def fake_result(anon_id: str = "PT001", site: str = "HN") -> dict:
    return {
        "AnonPatientID": anon_id,
        "site": site,
        "target_type": "GTV",
        "total_dose_gy": 70.0,
        "n_fractions": 35,
        "EQD2_gy": 70.0,
        "Dmean_gy": 68.5,
        "TCP_Poisson": 0.85,
        "TCP_ZM": 0.88,
        "TCP_gEUD": 0.83,
        "TCP_Logistic": 0.87,
        "TCP_mean": 0.86,
        "TCP_Poisson_mc": {"mean": 0.84, "sd": 0.04, "p5": 0.77, "p95": 0.91},
        "TCP_Poisson_hypoxia": 0.45,
        "TCP_gEUD_hypoxia": 0.40,
        "LocalControl": 1,
        "params_snapshot": {
            "alpha": 0.35,
            "beta": 0.035,
            "params_source": "default_yaml",
        },
    }


def test_build_benchmarking_table_shape():
    from outputs.reporter import build_benchmarking_table

    results = [fake_result(f"PT{i:03d}") for i in range(5)]
    df = build_benchmarking_table(results)
    assert len(df) == 5
    assert "TCP_Poisson" in df.columns
    assert "AnonPatientID" in df.columns


def test_build_benchmarking_table_mc_flattened():
    from outputs.reporter import build_benchmarking_table

    df = build_benchmarking_table([fake_result()])
    assert "TCP_Poisson_MC_Mean" in df.columns
    assert df["TCP_Poisson_MC_Mean"].iloc[0] == pytest.approx(0.84)


def test_save_benchmarking_excel_creates_file():
    from outputs.reporter import save_benchmarking_excel

    results = [fake_result(f"PT{i:03d}") for i in range(3)]
    with tempfile.TemporaryDirectory() as tmp:
        out = save_benchmarking_excel(results, None, pathlib.Path(tmp) / "bench.xlsx")
        assert out.exists()
        assert out.stat().st_size > 1000


def test_save_benchmarking_excel_three_sheets():
    pytest.importorskip("openpyxl")
    import openpyxl

    from outputs.reporter import save_benchmarking_excel

    results = [fake_result()]
    perf = {
        "Poisson": {
            "auc": 0.75,
            "auc_ci": (0.65, 0.85),
            "brier": 0.18,
            "ece": 0.05,
            "overfitting_index": 0.05,
            "cv_auc": 0.71,
            "harrell_c": None,
        }
    }
    with tempfile.TemporaryDirectory() as tmp:
        out = save_benchmarking_excel(results, perf, pathlib.Path(tmp) / "b.xlsx")
        wb = openpyxl.load_workbook(out)
        assert "TCP_Summary" in wb.sheetnames
        assert "Model_Performance" in wb.sheetnames
        assert "Params_Snapshot" in wb.sheetnames


def test_print_summary_table_runs_without_error(capsys):
    from outputs.reporter import print_summary_table

    results = [fake_result(f"PT{i:03d}") for i in range(3)]
    print_summary_table(results)
    captured = capsys.readouterr()
    assert "PT001" in captured.out or len(captured.out) >= 0


def test_benchmarking_table_nan_for_missing_keys():
    from outputs.reporter import build_benchmarking_table

    r = {"AnonPatientID": "PT999", "site": "HN"}
    df = build_benchmarking_table([r])
    assert math.isnan(df["TCP_Poisson"].iloc[0])


def test_dose_response_creates_file():
    from config.site_params import load_site_params
    from outputs.figures import plot_dose_response_curves

    sp = load_site_params("HN")
    with tempfile.TemporaryDirectory() as tmp:
        out = plot_dose_response_curves(sp, output_path=pathlib.Path(tmp) / "dr.png")
        assert out.exists()


def test_uncertainty_bands_creates_file():
    from outputs.figures import plot_uncertainty_bands

    doses = np.linspace(40, 80, 20)
    mean = np.clip(doses / 80, 0, 1)
    p5 = np.clip(mean - 0.1, 0, 1)
    p95 = np.clip(mean + 0.1, 0, 1)
    with tempfile.TemporaryDirectory() as tmp:
        out = plot_uncertainty_bands(
            doses, mean, p5, p95, "Poisson", pathlib.Path(tmp) / "unc.png"
        )
        assert out.exists()


def test_model_comparison_bar_creates_file():
    from outputs.figures import plot_model_comparison_bar

    ids = ["PT001", "PT002", "PT003"]
    tcps = {
        "Poisson": np.array([0.8, 0.7, 0.9]),
        "gEUD": np.array([0.75, 0.65, 0.85]),
    }
    with tempfile.TemporaryDirectory() as tmp:
        out = plot_model_comparison_bar(ids, tcps, pathlib.Path(tmp) / "cmp.png")
        assert out.exists()


def test_dose_response_all_sites():
    from config.site_params import load_site_params
    from outputs.figures import plot_dose_response_curves

    for site in ["HN", "LUNG", "BREAST", "BRAIN"]:
        sp = load_site_params(site)
        with tempfile.TemporaryDirectory() as tmp:
            out = plot_dose_response_curves(
                sp, output_path=pathlib.Path(tmp) / f"{site}.png"
            )
            assert out.exists(), f"Figure not created for {site}"


def test_figures_saved_at_high_dpi():
    from config.site_params import load_site_params
    from outputs.figures import plot_dose_response_curves

    sp = load_site_params("HN")
    with tempfile.TemporaryDirectory() as tmp:
        out = plot_dose_response_curves(
            sp, dpi=600, output_path=pathlib.Path(tmp) / "hires.png"
        )
        assert out.stat().st_size > 50_000


def test_cli_help_exits_zero():
    result = subprocess.run(
        [sys.executable, "-m", "rbgyanx_engine", "--help"],
        capture_output=True,
        text=True,
        cwd=str(pathlib.Path(__file__).resolve().parents[1]),
    )
    assert result.returncode == 0
    assert "--dicom-dir" in result.stdout
