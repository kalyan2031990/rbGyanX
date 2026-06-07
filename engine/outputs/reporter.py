# outputs/reporter.py

from __future__ import annotations

import pathlib
from typing import Any

import numpy as np
import pandas as pd


def build_benchmarking_table(
    cohort_results: list[dict],
) -> pd.DataFrame:
    """
    Convert list of per-patient result dicts into a single summary DataFrame.
    """
    rows = []
    for r in cohort_results:
        row = {
            "AnonPatientID": r.get("AnonPatientID", ""),
            "Site": r.get("site", ""),
            "Target": r.get("target_type", ""),
            "TotalDose_Gy": r.get("total_dose_gy", float("nan")),
            "NFractions": r.get("n_fractions", float("nan")),
            "EQD2_Gy": r.get("EQD2_gy", float("nan")),
            "Dmean_Gy": r.get("Dmean_gy", float("nan")),
            "TCP_Poisson": r.get("TCP_Poisson", float("nan")),
            "TCP_ZM": r.get("TCP_ZM", float("nan")),
            "TCP_gEUD": r.get("TCP_gEUD", float("nan")),
            "TCP_Logistic": r.get("TCP_Logistic", float("nan")),
            "TCP_Mean_Classical": r.get("TCP_mean", float("nan")),
            "TCP_Poisson_MC_Mean": r.get("TCP_Poisson_mc", {}).get("mean", float("nan")),
            "TCP_Poisson_MC_P5": r.get("TCP_Poisson_mc", {}).get("p5", float("nan")),
            "TCP_Poisson_MC_P95": r.get("TCP_Poisson_mc", {}).get("p95", float("nan")),
            "TCP_Poisson_Hypoxia": r.get("TCP_Poisson_hypoxia", float("nan")),
            "TCP_gEUD_Hypoxia": r.get("TCP_gEUD_hypoxia", float("nan")),
            "TCP_MVL": r.get("TCP_MVL", float("nan")),
            "TCP_XGBoost": r.get("TCP_XGBoost", float("nan")),
            "TCP_RF": r.get("TCP_RF", float("nan")),
            "UTCP": r.get("UTCP", float("nan")),
            "UTCP_weighted": r.get("UTCP_weighted", float("nan")),
            "UTCP_NTCP_product": r.get("UTCP_NTCP_product", float("nan")),
            "UTCP_OARs_scored": r.get("UTCP_OARs_scored", float("nan")),
            "UTCP_OARs_missing": r.get("UTCP_OARs_missing", float("nan")),
            "UTCP_missing_OARs": r.get("UTCP_missing_OARs", ""),
            "UTCP_TCP_model": r.get("UTCP_TCP_model", ""),
            "LocalControl": r.get("LocalControl", float("nan")),
            "ParamsSource": r.get("params_snapshot", {}).get("params_source", ""),
            "Data_Source": (
                "SYNTHETIC"
                if str(r.get("AnonPatientID", "")).startswith("AUG")
                else "CLINICAL"
            ),
        }
        rows.append(row)
    df = pd.DataFrame(rows)
    if "TCP_ZM" in df.columns:
        zm = df["TCP_ZM"]
        df["ZM_note"] = np.where(
            zm.notna(),
            "Poisson-N_eff approximation (see Zaider & Minerbo 2000)",
            "",
        )
    return df


def save_benchmarking_excel(
    cohort_results: list[dict],
    performance_metrics: dict | None,
    output_path: str | pathlib.Path,
) -> pathlib.Path:
    """
    Save multi-sheet benchmarking Excel workbook.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise ImportError("openpyxl required for Excel output.") from exc

    summary_df = build_benchmarking_table(cohort_results)

    params_rows = []
    for r in cohort_results:
        snap = r.get("params_snapshot", {})
        params_rows.append(
            {
                "AnonPatientID": r.get("AnonPatientID", ""),
                "Site": r.get("site", ""),
                **{k: v for k, v in snap.items()},
            }
        )
    params_df = pd.DataFrame(params_rows) if params_rows else pd.DataFrame()

    if performance_metrics:
        perf_rows = []
        for model_name, m in performance_metrics.items():
            if model_name.startswith("_") or not isinstance(m, dict):
                continue
            perf_rows.append(
                {
                    "Model": model_name,
                    "AUC": m.get("auc", float("nan")),
                    "AUC_CI_Lower": m.get("auc_ci", (float("nan"), float("nan")))[0],
                    "AUC_CI_Upper": m.get("auc_ci", (float("nan"), float("nan")))[1],
                    "Brier": m.get("brier", float("nan")),
                    "ECE": m.get("ece", float("nan")),
                    "Overfitting_Index": m.get("overfitting_index", float("nan")),
                    "CV_AUC": m.get("cv_auc", float("nan")),
                    "Harrell_C": m.get("harrell_c", float("nan")),
                    "Safety_Status": m.get("safety_status", ""),
                    "Safety_Annotation": m.get("safety_annotation", ""),
                }
            )
        perf_df = pd.DataFrame(perf_rows)
    else:
        perf_df = pd.DataFrame()

    out = pathlib.Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="TCP_Summary", index=False)
        perf_df.to_excel(writer, sheet_name="Model_Performance", index=False)
        params_df.to_excel(writer, sheet_name="Params_Snapshot", index=False)

        from openpyxl.styles import PatternFill

        orange_fill = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )
        red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        yellow_fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="99FF99", end_color="99FF99", fill_type="solid"
        )

        ws_tcp = writer.sheets["TCP_Summary"]
        if "Data_Source" in summary_df.columns:
            ds_col = summary_df.columns.get_loc("Data_Source") + 1
            for row_idx in range(2, len(summary_df) + 2):
                if ws_tcp.cell(row=row_idx, column=ds_col).value == "SYNTHETIC":
                    for col_idx in range(1, ws_tcp.max_column + 1):
                        ws_tcp.cell(row=row_idx, column=col_idx).fill = orange_fill

        if not perf_df.empty and "Safety_Status" in perf_df.columns:
            ws_perf = writer.sheets["Model_Performance"]
            status_col = perf_df.columns.get_loc("Safety_Status") + 1
            for row_idx in range(2, len(perf_df) + 2):
                status = ws_perf.cell(row=row_idx, column=status_col).value
                fill = None
                if status == "FAIL":
                    fill = red_fill
                elif status == "WARN":
                    fill = yellow_fill
                elif status == "PASS":
                    fill = green_fill
                if fill:
                    for col_idx in range(1, ws_perf.max_column + 1):
                        ws_perf.cell(row=row_idx, column=col_idx).fill = fill

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=8,
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    return out


def print_summary_table(
    cohort_results: list[dict],
    max_rows: int = 20,
) -> None:
    """
    Print a compact benchmarking summary to stdout using tabulate.
    """
    try:
        from tabulate import tabulate
    except ImportError:
        df = build_benchmarking_table(cohort_results)
        cols = [
            "AnonPatientID",
            "Site",
            "TCP_Poisson",
            "TCP_ZM",
            "TCP_gEUD",
            "TCP_Logistic",
        ]
        print(df[[c for c in cols if c in df.columns]].head(max_rows).to_string(index=False))
        return

    df = build_benchmarking_table(cohort_results).head(max_rows)
    cols = [
        "AnonPatientID",
        "Site",
        "TCP_Poisson",
        "TCP_ZM",
        "TCP_gEUD",
        "TCP_Logistic",
        "TCP_Poisson_Hypoxia",
    ]
    sub = df[[c for c in cols if c in df.columns]]
    float_cols = sub.select_dtypes(include="float").columns
    sub = sub.copy()
    sub[float_cols] = sub[float_cols].round(4)
    print(tabulate(sub, headers="keys", tablefmt="psql", showindex=False))
