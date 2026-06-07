"""
Clinical Data Template Generator for rbGyanX
=============================================
Generates Excel templates for TCP and NTCP analyses with sample data

Author: rbGyanX Team
Version: 1.0.0
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Try to import openpyxl for Excel formatting
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel formatting will be limited.")


class ClinicalTemplateGenerator:
    """
    Generate clinical data templates with documentation and sample data
    
    Supports:
    - TCP analysis templates (tumor control outcomes)
    - NTCP analysis templates (toxicity outcomes)
    - Combined TCP+NTCP templates
    - Multi-organ data structures
    - Sample synthetic data generation
    """
    
    def __init__(self):
        """Initialize template generator"""
        self.templates = {
            'TCP': self._tcp_template_schema(),
            'NTCP': self._ntcp_template_schema(),
            'TCP_NTCP': self._combined_template_schema()
        }
    
    def _tcp_template_schema(self):
        """TCP clinical data template schema"""
        return {
            'required_columns': [
                'PatientID',
                'TumorControl'  # 0=failure, 1=control
            ],
            'optional_columns': [
                'Age',
                'Gender',  # M/F/Other
                'Stage',  # I/II/III/IV
                'Histology',  # SCC/AdCa/Other
                'TumorVolume_cc',
                'TumorSite',  # Oropharynx/Nasopharynx/Larynx/etc
                'HPV_Status',  # Positive/Negative/Unknown
                'Smoking_PackYears',
                'ECOG_Performance',  # 0/1/2/3/4
                'ChemotherapyUsed',  # 0=no, 1=yes
                'Induction_Chemo',  # 0=no, 1=yes
                'Concurrent_Chemo',  # 0=no, 1=yes
                'TotalDose_Gy',
                'Fractions',
                'FractionationRegimen',  # Conventional/Hypofractionated/Hyperfractionated
                'TreatmentTechnique',  # IMRT/VMAT/3DCRT/Tomotherapy/Proton
                'FollowUp_Months',
                'LocalFailure_Time_Months',  # Time to failure (if applicable)
                'Notes'
            ],
            'column_descriptions': {
                'PatientID': 'Unique patient identifier (must match DVH CSV filenames)',
                'TumorControl': 'Local tumor control status: 0=failure/recurrence, 1=controlled',
                'Age': 'Patient age at treatment in years',
                'Gender': 'Patient gender: M=Male, F=Female, Other',
                'Stage': 'Clinical/pathological stage: I, II, III, IVA, IVB',
                'Histology': 'Tumor histology: SCC=Squamous Cell Carcinoma, AdCa=Adenocarcinoma',
                'TumorVolume_cc': 'Gross tumor volume in cubic centimeters',
                'TumorSite': 'Primary tumor site (e.g., Oropharynx, Nasopharynx)',
                'HPV_Status': 'HPV status for oropharyngeal cancers: Positive/Negative/Unknown',
                'Smoking_PackYears': 'Smoking history in pack-years',
                'ECOG_Performance': 'ECOG performance status: 0=fully active to 4=bedridden',
                'ChemotherapyUsed': 'Concurrent chemotherapy: 0=no, 1=yes',
                'TotalDose_Gy': 'Total prescribed dose in Gray',
                'Fractions': 'Number of treatment fractions',
                'TreatmentTechnique': 'Radiation delivery technique used',
                'FollowUp_Months': 'Follow-up duration in months',
                'LocalFailure_Time_Months': 'Time to local failure in months (if TumorControl=0)'
            },
            'data_types': {
                'PatientID': 'string',
                'TumorControl': 'binary (0/1)',
                'Age': 'numeric (30-90)',
                'Gender': 'categorical (M/F/Other)',
                'Stage': 'categorical (I/II/III/IV)',
                'TumorVolume_cc': 'numeric (>0)',
                'Smoking_PackYears': 'numeric (>=0)',
                'ChemotherapyUsed': 'binary (0/1)'
            }
        }
    
    def _ntcp_template_schema(self):
        """NTCP clinical data template schema"""
        return {
            'required_columns': [
                'PatientID',
                'Organ',  # OAR name
                'Toxicity'  # Primary toxicity endpoint (0/1 or grade 0-4)
            ],
            'optional_columns': [
                'Xerostomia_Binary',  # 0=no, 1=yes
                'Xerostomia_Grade',  # RTOG grade 0-4
                'Dysphagia_Binary',
                'Dysphagia_Grade',
                'Mucositis_Binary',
                'Mucositis_Grade',
                'Dermatitis_Binary',
                'Dermatitis_Grade',
                'Esophagitis_Binary',
                'Esophagitis_Grade',
                'Age',
                'Gender',
                'Stage',
                'Smoking_Status',  # Current/Former/Never
                'Diabetes',  # 0=no, 1=yes
                'ChemotherapyUsed',  # 0=no, 1=yes
                'Concurrent_Chemo',  # 0=no, 1=yes
                'TotalDose_Gy',
                'Fractions',
                'TreatmentTechnique',
                'BaselineSalivaryFlow_mL_min',  # For xerostomia
                'BaselineSwalloFunction_Score',  # For dysphagia
                'ToxicityAssessment_TimePoint',  # Acute (during RT) / Late (3-12 months)
                'FollowUp_Months',
                'Notes'
            ],
            'column_descriptions': {
                'PatientID': 'Unique patient identifier (must match DVH CSV filenames)',
                'Organ': 'Organ at risk name (must match organ in DVH filename, e.g., Parotid_L, Parotid_R)',
                'Toxicity': 'Primary toxicity endpoint: 0=no toxicity, 1=toxicity (or grade 0-4)',
                'Xerostomia_Binary': 'Xerostomia presence: 0=no, 1=yes (grade ≥2)',
                'Xerostomia_Grade': 'RTOG xerostomia grade: 0=none, 1=mild, 2=moderate, 3=severe, 4=life-threatening',
                'Dysphagia_Binary': 'Dysphagia presence: 0=no, 1=yes (grade ≥2)',
                'Dysphagia_Grade': 'RTOG dysphagia grade: 0-4',
                'Age': 'Patient age at treatment in years',
                'Gender': 'Patient gender: M=Male, F=Female, Other',
                'Smoking_Status': 'Smoking status: Current/Former/Never',
                'Diabetes': 'Diabetes diagnosis: 0=no, 1=yes',
                'ChemotherapyUsed': 'Concurrent chemotherapy: 0=no, 1=yes',
                'BaselineSalivaryFlow_mL_min': 'Baseline salivary flow rate (for xerostomia studies)',
                'ToxicityAssessment_TimePoint': 'When toxicity was assessed: Acute/Late',
                'FollowUp_Months': 'Follow-up duration in months'
            },
            'data_types': {
                'PatientID': 'string',
                'Organ': 'string',
                'Toxicity': 'binary (0/1) or grade (0-4)',
                'Age': 'numeric (30-90)',
                'Gender': 'categorical (M/F/Other)',
                'Smoking_Status': 'categorical',
                'Diabetes': 'binary (0/1)',
                'ChemotherapyUsed': 'binary (0/1)'
            }
        }
    
    def _combined_template_schema(self):
        """Combined TCP+NTCP template (merged columns)"""
        tcp_schema = self._tcp_template_schema()
        ntcp_schema = self._ntcp_template_schema()
        
        # Merge required columns
        required = list(set(tcp_schema['required_columns'] + ntcp_schema['required_columns']))
        
        # Merge optional columns (remove duplicates)
        optional = list(set(tcp_schema['optional_columns'] + ntcp_schema['optional_columns']))
        
        # Merge descriptions
        descriptions = {**tcp_schema['column_descriptions'], **ntcp_schema['column_descriptions']}
        
        return {
            'required_columns': required,
            'optional_columns': optional,
            'column_descriptions': descriptions,
            'data_types': {**tcp_schema['data_types'], **ntcp_schema['data_types']}
        }
    
    def generate_template(self, template_type='TCP', output_path=None, 
                         include_sample_data=True, n_samples=30):
        """
        Generate clinical data template Excel file
        
        Parameters
        ----------
        template_type : str, default 'TCP'
            Template type: 'TCP', 'NTCP', or 'TCP_NTCP'
        output_path : Path or str, optional
            Output file path. If None, auto-generates filename
        include_sample_data : bool, default True
            If True, populate with synthetic sample data
        n_samples : int, default 30
            Number of sample patients to generate
        
        Returns
        -------
        Path
            Path to generated template file
        """
        if template_type not in self.templates:
            raise ValueError(f"Invalid template type: {template_type}. Must be TCP, NTCP, or TCP_NTCP")
        
        schema = self.templates[template_type]
        
        # Generate data
        if include_sample_data:
            print(f"Generating {n_samples} sample patients for {template_type} analysis...")
            df = self._generate_sample_data(schema, n_samples, template_type)
        else:
            # Empty template with column headers only
            columns = schema['required_columns'] + schema['optional_columns']
            df = pd.DataFrame(columns=columns)
        
        # Set output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(f"rbGyanX_Clinical_Template_{template_type}_{timestamp}.xlsx")
        else:
            output_path = Path(output_path)
        
        # Create parent directory if needed
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create Excel workbook with multiple sheets
        print(f"Creating Excel template: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Clinical Data
            df.to_excel(writer, sheet_name='ClinicalData', index=False)
            
            # Sheet 2: Column Descriptions
            descriptions_df = self._create_descriptions_sheet(schema)
            descriptions_df.to_excel(writer, sheet_name='ColumnDescriptions', index=False)
            
            # Sheet 3: Instructions
            instructions_df = self._create_instructions_sheet(template_type, schema)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format workbook if openpyxl available
            if OPENPYXL_AVAILABLE:
                self._format_excel_workbook(writer, df, descriptions_df, instructions_df)
        
        print(f"[OK] Template created successfully: {output_path}")
        print(f"     - {len(df)} sample patients")
        print(f"     - {len(df.columns)} columns")
        print(f"     - 3 sheets (Data, Descriptions, Instructions)")
        
        return output_path
    
    def _generate_sample_data(self, schema, n_samples, template_type):
        """Generate realistic synthetic clinical data"""
        np.random.seed(42)  # Reproducible sample data
        
        data = {}
        
        # Required columns
        if 'PatientID' in schema['required_columns']:
            data['PatientID'] = [f"Patient_{i:03d}" for i in range(n_samples)]
        
        if 'TumorControl' in schema['required_columns']:
            # 70% tumor control rate (typical for HNSCC)
            data['TumorControl'] = np.random.binomial(1, 0.70, n_samples)
        
        if 'Organ' in schema['required_columns']:
            # For NTCP: alternate between organs (bilateral parotids typical)
            organs = []
            for i in range(n_samples):
                if i % 2 == 0:
                    organs.append('Parotid_L')
                else:
                    organs.append('Parotid_R')
            data['Organ'] = organs
        
        if 'Toxicity' in schema['required_columns']:
            # 40% toxicity rate (grade ≥2) - typical for xerostomia
            data['Toxicity'] = np.random.binomial(1, 0.40, n_samples)
        
        # Optional columns - demographics
        if 'Age' in schema['optional_columns']:
            # Mean age 62, SD 12 years (typical for head-neck cancer)
            data['Age'] = np.random.normal(62, 12, n_samples).astype(int)
            data['Age'] = np.clip(data['Age'], 30, 90)
        
        if 'Gender' in schema['optional_columns']:
            # 65% male (typical for head-neck cancer)
            data['Gender'] = np.random.choice(['M', 'F'], n_samples, p=[0.65, 0.35])
        
        if 'Stage' in schema['optional_columns']:
            # Stage distribution (more advanced stages common)
            data['Stage'] = np.random.choice(['I', 'II', 'III', 'IVA', 'IVB'], 
                                            n_samples, p=[0.05, 0.15, 0.25, 0.35, 0.20])
        
        # Tumor characteristics
        if 'TumorVolume_cc' in schema['optional_columns']:
            # Log-normal distribution, median ~45 cc
            data['TumorVolume_cc'] = np.random.lognormal(3.8, 0.7, n_samples)
            data['TumorVolume_cc'] = np.round(data['TumorVolume_cc'], 1)
        
        if 'TumorSite' in schema['optional_columns']:
            sites = ['Oropharynx', 'Larynx', 'Hypopharynx', 'Nasopharynx']
            data['TumorSite'] = np.random.choice(sites, n_samples, 
                                                p=[0.45, 0.30, 0.15, 0.10])
        
        if 'Histology' in schema['optional_columns']:
            # Mostly SCC in head-neck
            data['Histology'] = np.random.choice(['SCC', 'AdCa', 'Other'], 
                                                 n_samples, p=[0.85, 0.10, 0.05])
        
        if 'HPV_Status' in schema['optional_columns']:
            # For oropharynx, ~60% HPV+ nowadays
            hpv = []
            for site in data.get('TumorSite', ['Oropharynx']*n_samples):
                if site == 'Oropharynx':
                    hpv.append(np.random.choice(['Positive', 'Negative', 'Unknown'], 
                                               p=[0.60, 0.30, 0.10]))
                else:
                    hpv.append(np.random.choice(['Negative', 'Unknown'], p=[0.80, 0.20]))
            data['HPV_Status'] = hpv
        
        # Smoking
        if 'Smoking_PackYears' in schema['optional_columns']:
            # ~60% smokers, mean 30 pack-years
            smoking = []
            for _ in range(n_samples):
                if np.random.random() < 0.60:
                    smoking.append(np.round(np.random.gamma(3, 10), 1))
                else:
                    smoking.append(0)
            data['Smoking_PackYears'] = smoking
        
        if 'Smoking_Status' in schema['optional_columns']:
            smoking_status = []
            for py in data.get('Smoking_PackYears', [0]*n_samples):
                if py == 0:
                    smoking_status.append('Never')
                elif np.random.random() < 0.40:
                    smoking_status.append('Former')
                else:
                    smoking_status.append('Current')
            data['Smoking_Status'] = smoking_status
        
        # Performance status
        if 'ECOG_Performance' in schema['optional_columns']:
            data['ECOG_Performance'] = np.random.choice([0, 1, 2], n_samples, 
                                                        p=[0.50, 0.35, 0.15])
        
        # Treatment parameters
        if 'TotalDose_Gy' in schema['optional_columns']:
            # Standard fractionation: 70 Gy most common
            doses = np.random.choice([66, 70, 72, 74], n_samples, 
                                    p=[0.15, 0.60, 0.15, 0.10])
            data['TotalDose_Gy'] = doses
        
        if 'Fractions' in schema['optional_columns']:
            # Mostly conventional 2 Gy fractions
            fractions = []
            for dose in data.get('TotalDose_Gy', [70]*n_samples):
                if dose == 70:
                    fractions.append(35)
                elif dose == 66:
                    fractions.append(33)
                elif dose == 72:
                    fractions.append(36)
                else:
                    fractions.append(37)
            data['Fractions'] = fractions
        
        if 'FractionationRegimen' in schema['optional_columns']:
            data['FractionationRegimen'] = ['Conventional'] * n_samples
        
        if 'TreatmentTechnique' in schema['optional_columns']:
            techniques = np.random.choice(['IMRT', 'VMAT', 'Tomotherapy'], 
                                         n_samples, p=[0.40, 0.50, 0.10])
            data['TreatmentTechnique'] = techniques
        
        if 'ChemotherapyUsed' in schema['optional_columns']:
            # ~70% get concurrent chemo for advanced stage
            chemo = []
            for stage in data.get('Stage', ['III']*n_samples):
                if stage in ['III', 'IVA', 'IVB']:
                    chemo.append(np.random.binomial(1, 0.70))
                else:
                    chemo.append(np.random.binomial(1, 0.30))
            data['ChemotherapyUsed'] = chemo
        
        if 'Concurrent_Chemo' in schema['optional_columns']:
            data['Concurrent_Chemo'] = data.get('ChemotherapyUsed', [0]*n_samples)
        
        # Toxicity grades
        if 'Xerostomia_Grade' in schema['optional_columns']:
            grades = []
            for tox in data.get('Toxicity', [0]*n_samples):
                if tox == 1:
                    grades.append(np.random.choice([2, 3, 4], p=[0.60, 0.30, 0.10]))
                else:
                    grades.append(np.random.choice([0, 1], p=[0.70, 0.30]))
            data['Xerostomia_Grade'] = grades
        
        if 'Xerostomia_Binary' in schema['optional_columns']:
            # Grade ≥2 = clinically significant
            if 'Xerostomia_Grade' in data:
                data['Xerostomia_Binary'] = [1 if g >= 2 else 0 for g in data['Xerostomia_Grade']]
            else:
                data['Xerostomia_Binary'] = data.get('Toxicity', [0]*n_samples)
        
        # Follow-up
        if 'FollowUp_Months' in schema['optional_columns']:
            # Median 24 months, range 6-60
            data['FollowUp_Months'] = np.random.gamma(4, 6, n_samples).astype(int)
            data['FollowUp_Months'] = np.clip(data['FollowUp_Months'], 6, 60)
        
        # Other optional columns with defaults
        if 'Diabetes' in schema['optional_columns']:
            data['Diabetes'] = np.random.binomial(1, 0.20, n_samples)
        
        if 'ToxicityAssessment_TimePoint' in schema['optional_columns']:
            data['ToxicityAssessment_TimePoint'] = np.random.choice(
                ['Acute', 'Late'], n_samples, p=[0.30, 0.70])
        
        if 'Notes' in schema['optional_columns']:
            data['Notes'] = ['Sample data - replace with actual clinical information'] * n_samples
        
        return pd.DataFrame(data)
    
    def _create_descriptions_sheet(self, schema):
        """Create column descriptions reference sheet"""
        descriptions = []
        
        all_columns = schema['required_columns'] + schema['optional_columns']
        
        for col in all_columns:
            desc = schema['column_descriptions'].get(col, 'No description available')
            dtype = schema['data_types'].get(col, 'Not specified')
            required = 'REQUIRED' if col in schema['required_columns'] else 'Optional'
            
            descriptions.append({
                'Column Name': col,
                'Required/Optional': required,
                'Data Type': dtype,
                'Description': desc
            })
        
        return pd.DataFrame(descriptions)
    
    def _create_instructions_sheet(self, template_type, schema):
        """Create instructions sheet"""
        instructions = []
        
        instructions.append({
            'Section': 'TEMPLATE TYPE',
            'Information': f'{template_type} Analysis Clinical Data Template'
        })
        
        instructions.append({
            'Section': 'PURPOSE',
            'Information': f'This template is for {template_type} analysis in rbGyanX BASIC'
        })
        
        instructions.append({
            'Section': 'REQUIRED COLUMNS',
            'Information': ', '.join(schema['required_columns'])
        })
        
        instructions.append({
            'Section': 'OPTIONAL COLUMNS',
            'Information': ', '.join(schema['optional_columns'])
        })
        
        instructions.append({
            'Section': 'PATIENT ID MATCHING',
            'Information': 'PatientID must exactly match the patient identifier in DVH CSV filenames (e.g., Patient_001 matches Patient_001_Parotid.csv)'
        })
        
        instructions.append({
            'Section': 'SAMPLE DATA',
            'Information': 'This template includes 30 rows of synthetic sample data. You can either modify this data or delete all rows and enter your own clinical data.'
        })
        
        instructions.append({
            'Section': 'MINIMUM REQUIREMENTS',
            'Information': 'For meaningful analysis: ≥15 patients total, ≥5 events (TumorControl=1 or Toxicity=1) for ML models'
        })
        
        instructions.append({
            'Section': 'MULTI-ORGAN DATA (NTCP)',
            'Information': 'For NTCP analysis of multiple organs per patient, create one row per organ (e.g., Patient_001 Parotid_L, Patient_001 Parotid_R)'
        })
        
        instructions.append({
            'Section': 'MISSING DATA',
            'Information': 'Leave cells blank for missing optional data. Required columns must not have missing values.'
        })
        
        instructions.append({
            'Section': 'NEXT STEPS',
            'Information': '1. Fill in/modify clinical data in ClinicalData sheet. 2. Save file. 3. In rbGyanX GUI, select this file as Clinical Data input. 4. Run analysis.'
        })
        
        return pd.DataFrame(instructions)
    
    def _format_excel_workbook(self, writer, data_df, descriptions_df, instructions_df):
        """Apply Excel formatting for professional appearance"""
        if not OPENPYXL_AVAILABLE:
            return
        
        workbook = writer.book
        
        # Format ClinicalData sheet
        data_sheet = writer.sheets['ClinicalData']
        self._format_data_sheet(data_sheet, data_df)
        
        # Format ColumnDescriptions sheet
        desc_sheet = writer.sheets['ColumnDescriptions']
        self._format_descriptions_sheet(desc_sheet, descriptions_df)
        
        # Format Instructions sheet
        inst_sheet = writer.sheets['Instructions']
        self._format_instructions_sheet(inst_sheet, instructions_df)
    
    def _format_data_sheet(self, sheet, df):
        """Format clinical data sheet"""
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 3, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                   min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
    
    def _format_descriptions_sheet(self, sheet, df):
        """Format column descriptions sheet"""
        # Similar formatting to data sheet
        self._format_data_sheet(sheet, df)
        
        # Additional: Wrap text in Description column
        for cell in sheet['D']:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Wider Description column
        sheet.column_dimensions['D'].width = 60
    
    def _format_instructions_sheet(self, sheet, df):
        """Format instructions sheet"""
        # Similar to descriptions
        self._format_data_sheet(sheet, df)
        
        # Wrap text
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Wider Information column
        sheet.column_dimensions['B'].width = 80


# Convenience functions for quick template generation
def create_tcp_template(output_path=None, with_samples=True, n_samples=30):
    """
    Quick function to create TCP clinical data template
    
    Parameters
    ----------
    output_path : Path or str, optional
        Where to save template
    with_samples : bool, default True
        Include sample data
    n_samples : int, default 30
        Number of sample patients
    
    Returns
    -------
    Path
        Path to created template
    """
    generator = ClinicalTemplateGenerator()
    return generator.generate_template('TCP', output_path, with_samples, n_samples)


def create_ntcp_template(output_path=None, with_samples=True, n_samples=30):
    """
    Quick function to create NTCP clinical data template
    
    Parameters
    ----------
    output_path : Path or str, optional
        Where to save template
    with_samples : bool, default True
        Include sample data
    n_samples : int, default 30
        Number of sample patients
    
    Returns
    -------
    Path
        Path to created template
    """
    generator = ClinicalTemplateGenerator()
    return generator.generate_template('NTCP', output_path, with_samples, n_samples)


def create_combined_template(output_path=None, with_samples=True, n_samples=30):
    """
    Quick function to create combined TCP+NTCP template
    
    Parameters
    ----------
    output_path : Path or str, optional
        Where to save template
    with_samples : bool, default True
        Include sample data
    n_samples : int, default 30
        Number of sample patients
    
    Returns
    -------
    Path
        Path to created template
    """
    generator = ClinicalTemplateGenerator()
    return generator.generate_template('TCP_NTCP', output_path, with_samples, n_samples)


# Command-line interface
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate clinical data templates for rbGyanX',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Create TCP template
  python clinical_template_generator.py --type TCP --output tcp_template.xlsx
  
  # Create NTCP template without sample data
  python clinical_template_generator.py --type NTCP --no-samples
  
  # Create combined template with 50 samples
  python clinical_template_generator.py --type TCP_NTCP --samples 50
        """
    )
    
    parser.add_argument('--type', choices=['TCP', 'NTCP', 'TCP_NTCP'], 
                       default='TCP',
                       help='Template type (default: TCP)')
    parser.add_argument('--output', type=str, default=None,
                       help='Output file path (default: auto-generated)')
    parser.add_argument('--samples', type=int, default=30,
                       help='Number of sample patients (default: 30)')
    parser.add_argument('--no-samples', action='store_true',
                       help='Create empty template without sample data')
    
    args = parser.parse_args()
    
    # Generate template
    generator = ClinicalTemplateGenerator()
    template_path = generator.generate_template(
        template_type=args.type,
        output_path=args.output,
        include_sample_data=not args.no_samples,
        n_samples=args.samples
    )
    
    print(f"\n{'='*60}")
    print(f"Template ready: {template_path}")
    print(f"{'='*60}")
    print("\nNext steps:")
    print("1. Open the template in Excel")
    print("2. Review sample data and column descriptions")
    print("3. Replace sample data with your clinical information")
    print("4. Save the file")
    print("5. Use in rbGyanX GUI as Clinical Data input")
